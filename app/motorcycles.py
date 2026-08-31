from datetime import date
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, url_for, send_file
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app import db
from app.models import City, ImmobilizedMotorcycle

motorcycles_bp = Blueprint(
    "motorcycles", __name__, url_prefix="/motos-imobilizadas"
)

STATUS_OPTIONS = (
    "Solicitações de garantia",
    "Aguardando chegada das peças",
    "Aguardando aprovação do seguro",
    "Aguardando diagnóstico",
    "Aguardando peças",
    "Em manutenção",
    "Pronta para retirada",
    "Finalizada",
)


def _accessible_query():
    query = ImmobilizedMotorcycle.query
    if not current_user.is_admin:
        query = query.filter_by(city_id=current_user.city_id)
    return query


def _get_accessible_or_404(motorcycle_id):
    motorcycle = _accessible_query().filter_by(id=motorcycle_id).first()
    if motorcycle is None:
        abort(404)
    return motorcycle


def _city_from_form():
    if not current_user.is_admin:
        return current_user.city
    try:
        city_id = int(request.form.get("city_id", ""))
    except ValueError:
        return None
    return db.session.get(City, city_id)


def _form_values():
    try:
        entry_date = date.fromisoformat(request.form.get("entry_date", ""))
        expected_date = date.fromisoformat(request.form.get("expected_date", ""))
    except ValueError:
        entry_date = expected_date = None

    return {
        "city": _city_from_form(),
        "client": request.form.get("client", "").strip(),
        "model": request.form.get("model", "").strip(),
        "plate": request.form.get("plate", "").strip().upper(),
        "chassis": request.form.get("chassis", "").strip().upper(),
        "service_order": request.form.get("service_order", "").strip(),
        "reason": request.form.get("reason", "").strip(),
        "entry_date": entry_date,
        "expected_date": expected_date,
        "status": request.form.get("status", "").strip(),
        "responsible": request.form.get("responsible", "").strip() or None,
        "notes": request.form.get("notes", "").strip() or None,
    }


def _validation_error(values):
    if any(
        not values[field]
        for field in (
            "city",
            "client",
            "model",
            "plate",
            "chassis",
            "service_order",
            "reason",
            "entry_date",
            "expected_date",
            "status",
        )
    ):
        return "Preencha todos os campos obrigatórios."
    if values["status"] not in STATUS_OPTIONS:
        return "Selecione um status válido."
    if values["expected_date"] < values["entry_date"]:
        return "A previsão não pode ser anterior à data de entrada."
    return None


@motorcycles_bp.route("/")
@login_required
def index():
    motorcycles = _accessible_query().order_by(
        ImmobilizedMotorcycle.entry_date.desc(), ImmobilizedMotorcycle.id.desc()
    ).all()
    return render_template(
        "motorcycles/index.html", motorcycles=motorcycles, today=date.today()
    )


@motorcycles_bp.route("/nova", methods=["GET", "POST"])
@login_required
def create():
    cities = City.query.order_by(City.name).all() if current_user.is_admin else []
    if request.method == "POST":
        values = _form_values()
        error = _validation_error(values)
        if error:
            flash(error, "danger")
        else:
            motorcycle = ImmobilizedMotorcycle(**values)
            db.session.add(motorcycle)
            db.session.commit()
            flash("Moto imobilizada cadastrada com sucesso.", "success")
            return redirect(url_for("motorcycles.index"))
    return render_template(
        "motorcycles/form.html",
        motorcycle=None,
        cities=cities,
        statuses=STATUS_OPTIONS,
    )


@motorcycles_bp.route("/<int:motorcycle_id>/editar", methods=["GET", "POST"])
@login_required
def edit(motorcycle_id):
    motorcycle = _get_accessible_or_404(motorcycle_id)
    cities = City.query.order_by(City.name).all() if current_user.is_admin else []
    if request.method == "POST":
        values = _form_values()
        error = _validation_error(values)
        if error:
            flash(error, "danger")
        else:
            for field, value in values.items():
                setattr(motorcycle, field, value)
            db.session.commit()
            flash("Moto imobilizada atualizada com sucesso.", "success")
            return redirect(url_for("motorcycles.index"))
    return render_template(
        "motorcycles/form.html",
        motorcycle=motorcycle,
        cities=cities,
        statuses=STATUS_OPTIONS,
    )


@motorcycles_bp.route("/<int:motorcycle_id>/excluir", methods=["POST"])
@login_required
def delete(motorcycle_id):
    motorcycle = _get_accessible_or_404(motorcycle_id)
    db.session.delete(motorcycle)
    db.session.commit()
    flash("Moto imobilizada excluída com sucesso.", "success")
    return redirect(url_for("motorcycles.index"))


def _export_to_excel(motorcycles):
    """Gera um arquivo Excel com os dados das motos imobilizadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Motos Imobilizadas"
    
    # Definir largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 30
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    headers = [
        "ID", "Cidade", "Cliente", "Modelo", "Placa", "Chassi",
        "Ordem Serviço", "Motivo", "Data Entrada", "Previsão", "Status", "Responsável", "Observações"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Dados
    for row_num, motorcycle in enumerate(motorcycles, 2):
        data = [
            motorcycle.id,
            motorcycle.city.name if motorcycle.city else "",
            motorcycle.client,
            motorcycle.model,
            motorcycle.plate,
            motorcycle.chassis,
            motorcycle.service_order,
            motorcycle.reason,
            motorcycle.entry_date.isoformat() if motorcycle.entry_date else "",
            motorcycle.expected_date.isoformat() if motorcycle.expected_date else "",
            motorcycle.status,
            motorcycle.responsible or "",
            motorcycle.notes or "",
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Salvar em bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@motorcycles_bp.route("/exportar-excel", methods=["POST"])
@login_required
def export_excel():
    """Endpoint para fazer download da planilha em Excel."""
    motorcycles = _accessible_query().order_by(
        ImmobilizedMotorcycle.entry_date.desc(), ImmobilizedMotorcycle.id.desc()
    ).all()
    
    if not motorcycles:
        flash("Nenhuma moto para exportar.", "warning")
        return redirect(url_for("motorcycles.index"))
    
    excel_file = _export_to_excel(motorcycles)
    
    # Nome do arquivo com data
    from datetime import datetime
    filename = f"motos-imobilizadas-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
